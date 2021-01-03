#!/usr/bin/env python

from openpyxl import load_workbook

# ファイル名
excel_read_file_name = "kuku.xlsx"

# ワークブックオブジェクトの作成
# <class 'openpyxl.workbook.workbook.Workbook'>
wb = load_workbook(excel_read_file_name)

# シート名の一覧を取得
sheetname = wb.sheetnames

# 1つ目のシートを指定して、ワークシートオブジェクトを作成
# <class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet = wb[sheetname[0]]

# 最大行数の取得
max_row = sheet.max_row

# 最大列数の取得
max_col = sheet.max_column

# 取得データの表示
for i in range(1, max_row + 1):
  for j in range(1, max_col + 1):
    cell = sheet.cell(column=j, row=i).value
    print(str(cell).rjust(2, " ") + " ", end="")
  print("")